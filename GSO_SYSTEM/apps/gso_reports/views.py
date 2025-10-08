# apps/gso_reports/views.py
import os
import json
import calendar
import openpyxl
from datetime import datetime
from django.conf import settings
from django.http import HttpResponse
from django.db.models import Q

from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, JsonResponse

from apps.gso_requests.models import ServiceRequest
from apps.gso_accounts.models import User, Unit
from .models import WorkAccomplishmentReport, SuccessIndicator, IPMT, ActivityName
from .utils import normalize_report, generate_ipmt_excel, collect_ipmt_reports
from apps.ai_service.utils import generate_war_description, generate_ipmt_summary


# -------------------------------
# Role Checks
# -------------------------------
def is_gso_or_director(user):
    return user.is_authenticated and user.role in ["gso", "director"]


# -------------------------------
# Accomplishment Report View
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def accomplishment_report(request):
    # Fetch completed requests
    completed_requests = ServiceRequest.objects.filter(status="Completed").order_by("-created_at")
    # Fetch all WARs
    all_wars = WorkAccomplishmentReport.objects.select_related("request", "unit").prefetch_related("assigned_personnel").all().order_by("-date_started")

    reports = []

    # Track which requests already have a WAR
    war_request_ids = set(war.request_id for war in all_wars if war.request_id)

    # Process completed requests that **don't yet have a WAR**
    for r in completed_requests:
        if r.id in war_request_ids:
            continue
        norm = normalize_report(r)
        norm["id"] = r.id

        if not norm.get("description") or norm["description"].strip() == "":
            try:
                desc = generate_war_description(
                    activity_name=getattr(r, "activity_name", getattr(r, "title", "Task")),
                    unit=getattr(r.unit, "name", None),
                    personnel_names=[p.get_full_name() for p in r.assigned_personnel.all()] if hasattr(r, "assigned_personnel") else None
                )
                r.description = desc or "No description generated."
                r.save(update_fields=["description"])
                norm["description"] = r.description
            except Exception as e:
                norm["description"] = f"Error generating description: {e}"

        reports.append(norm)

    # Process all WARs
    for war in all_wars:
        norm = normalize_report(war)
        norm["id"] = war.id

        if not norm.get("description") or norm["description"].strip() == "":
            try:
                desc = generate_war_description(
                    activity_name=getattr(war, "activity_name", getattr(war, "title", "Task")),
                    unit=getattr(war.unit, "name", None),
                    personnel_names=[p.get_full_name() for p in war.assigned_personnel.all()] if hasattr(war, "assigned_personnel") else None
                )
                war.description = desc or "No description generated."
                war.save(update_fields=["description"])
                norm["description"] = war.description
            except Exception as e:
                norm["description"] = f"Error generating description: {e}"

        reports.append(norm)

    # Apply search and unit filters
    search_query = request.GET.get("q")
    if search_query:
        reports = [r for r in reports if search_query.lower() in str(r).lower()]

    unit_filter = request.GET.get("unit")
    if unit_filter:
        reports = [r for r in reports if r["unit"].lower() == unit_filter.lower()]

    reports.sort(key=lambda r: r["date"], reverse=True)

    # Load all active personnel for IPMT modal
    personnel_qs = User.objects.filter(role="personnel", account_status="active") \
        .select_related('unit').order_by('unit__name', 'first_name')

    personnel_list = [
        {
            "full_name": u.get_full_name() or u.username,
            "username": u.username,
            "unit": u.unit.name.lower() if u.unit else "unassigned"
        }
        for u in personnel_qs
    ]

    return render(
        request,
        "gso_office/accomplishment_report/accomplishment_report.html",
        {
            "reports": reports,
            "personnel_list": personnel_list,
        },
    )

# -------------------------------
# Generate IPMT Excel
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def generate_ipmt(request):
    import json
    import calendar
    import os
    import openpyxl
    from django.http import HttpResponse
    from django.conf import settings

    reports = []
    personnel_list = []

    # --- Handle POST (from edited preview OR export form) ---
    if request.method == "POST":
        # ✅ Handle both JSON and form POST
        try:
            # Try JSON (Save button)
            body = json.loads(request.body.decode("utf-8"))
            month_filter = body.get("month")
            unit_filter = body.get("unit")
            personnel_param = body.get("personnel", "")
            reports = body.get("rows", [])
        except Exception:
            # Fallback to form POST (Export button)
            month_filter = request.POST.get("month")
            unit_filter = request.POST.get("unit")
            personnel_param = request.POST.get("personnel", "")
            rows_data = request.POST.get("rows", "[]")
            try:
                reports = json.loads(rows_data)
            except json.JSONDecodeError:
                reports = []

        # ✅ Convert personnel list
        personnel_list = [p.strip() for p in personnel_param.split(",") if p.strip()]

        # Update indicator to include code + description from SuccessIndicator
        for r in reports:
            if not r.get("indicator"):
                continue
            code_only = r["indicator"].split(" - ")[0].strip()
            si = SuccessIndicator.objects.filter(code__iexact=code_only).first()
            if si:
                r["indicator"] = f"{si.code} - {si.description}"

    # --- Handle GET fallback (for debugging / direct access) ---
    else:
        personnel_param = request.GET.get("personnel")
        month_filter = request.GET.get("month")
        unit_filter = request.GET.get("unit")

        if not personnel_param or not month_filter or not unit_filter:
            return HttpResponse("Personnel, unit, and month are required.", status=400)

        year, month_num = map(int, month_filter.split("-"))
        month_name = f"{calendar.month_name[month_num]} {year}"
        personnel_list = [p.strip() for p in personnel_param.split(",")]

        for identifier in personnel_list:
            user = get_user_by_identifier(identifier)
            if not user:
                continue

            # --- Fetch saved IPMT rows first ---
            ipmt_rows = IPMT.objects.filter(
                personnel=user,
                unit__name__iexact=unit_filter,
                month=f"{calendar.month_name[month_num]} {year}"
            )

            if ipmt_rows.exists():
                for row in ipmt_rows:
                    reports.append({
                        "indicator": f"{row.indicator.code} - {row.indicator.description}",
                        "description": row.accomplishment or "",
                        "remarks": row.remarks or "",
                    })
            else:
                # --- Fallback: WARs ---
                wars = WorkAccomplishmentReport.objects.filter(
                    assigned_personnel=user,
                    unit__name__iexact=unit_filter,
                    date_started__year=year,
                    date_started__month=month_num
                )

                for war in wars:
                    activity_obj = None
                    if war.activity_name:
                        activity_obj = ActivityName.objects.filter(name__iexact=war.activity_name).first()

                    sis = SuccessIndicator.objects.filter(
                        unit=war.unit,
                        activity_name=activity_obj
                    )

                    if sis.exists():
                        for si in sis:
                            reports.append({
                                "indicator": f"{si.code} - {si.description}",
                                "description": war.description,
                                "remarks": "Complied" if war.description else "",
                            })
                    else:
                        reports.append({
                            "indicator": war.activity_name or war.unit.name,
                            "description": war.description,
                            "remarks": "Complied" if war.description else "",
                        })

                # --- Fallback: Completed ServiceRequests ---
                completed_requests = ServiceRequest.objects.filter(
                    assigned_personnel=user,
                    unit__name__iexact=unit_filter,
                    status="Completed",
                    created_at__year=year,
                    created_at__month=month_num
                )

                for req in completed_requests:
                    reports.append({
                        "indicator": req.unit.name,
                        "description": req.description,
                        "remarks": "Complied" if req.description else "",
                    })

    # --- Load Excel template ---
    template_path = os.path.join(settings.BASE_DIR, "static", "excel_file", "sampleipmt.xlsx")
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # --- Build personnel full names ---
    personnel_fullnames = []
    for identifier in personnel_list:
        user_obj = get_user_by_identifier(identifier)
        if user_obj:
            full_name = (user_obj.get_full_name() or "").strip()
            if full_name:
                personnel_fullnames.append(full_name)
            elif user_obj.first_name or user_obj.last_name:
                personnel_fullnames.append(f"{user_obj.first_name} {user_obj.last_name}".strip())
            else:
                personnel_fullnames.append(user_obj.username)
        else:
            personnel_fullnames.append(identifier)

    # --- Debug log ---
    print("=== IPMT Personnel Fullnames ===")
    print(personnel_fullnames)

    # --- Write to Excel ---
    try:
        ws["B8"] = ", ".join(personnel_fullnames) if personnel_fullnames else "No personnel found"
    except Exception as e:
        print(f"Error writing to Excel B8: {e}")
        ws.cell(row=8, column=2, value=", ".join(personnel_fullnames))

    # --- Write month ---
    try:
        if "-" in month_filter:
            year, month_num = map(int, month_filter.split("-"))
            month_name = f"{calendar.month_name[month_num]} {year}"
        else:
            month_name = month_filter
        ws["B11"] = month_name
    except Exception as e:
        print(f"Error writing to Excel B11: {e}")

    # --- Write reports ---
    start_row = 13
    for i, r in enumerate(reports, start=start_row):
        ws.cell(row=i, column=1).value = r.get("indicator", "")
        ws.cell(row=i, column=2).value = r.get("description", "")
        ws.cell(row=i, column=3).value = r.get("remarks", "")

    # --- Return Excel ---
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"IPMT_{unit_filter}_{month_filter}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# -------------------------------
# Get WAR Description (AJAX)
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def get_war_description(request, war_id):
    try:
        war = WorkAccomplishmentReport.objects.get(id=war_id)
        return JsonResponse({'description': war.description or ""})
    except WorkAccomplishmentReport.DoesNotExist:
        return JsonResponse({'error': 'WAR not found'}, status=404)
# -------------------------------
# Preview IPMT (Web)
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def preview_ipmt(request):
    """
    Preview IPMT rows for the selected unit, personnel, and month.
    Fetches WorkAccomplishmentReports matching each SuccessIndicator's activity_name.
    """
    month_filter = request.GET.get("month")
    unit_filter = request.GET.get("unit", "all")
    personnel_names = request.GET.getlist("personnel[]") or []

    if not month_filter:
        return HttpResponse("Month is required in 'YYYY-MM' format.", status=400)

    try:
        year, month_num = map(int, month_filter.split("-"))
    except ValueError:
        return HttpResponse("Invalid month format. Use YYYY-MM.", status=400)

    from apps.gso_accounts.models import User, Unit
    from .models import SuccessIndicator, WorkAccomplishmentReport

    unit = Unit.objects.filter(name__iexact=unit_filter).first()
    if not unit:
        return HttpResponse("Unit not found.", status=404)

    reports = []

    for person_name in personnel_names:
        # Lookup by full name (adjust as needed)
        user = get_user_by_identifier(person_name)
        if not user:
            continue

        # Get all active success indicators for this unit
        indicators = SuccessIndicator.objects.filter(unit=unit, is_active=True)

        for indicator in indicators:
            # Determine which activity_name to match against WARs
            activity_name_to_match = (
                indicator.activity_name.name if indicator.activity_name else indicator.code
            )

            # Fetch related WARs for this user and activity_name
            wars = WorkAccomplishmentReport.objects.filter(
                unit=unit,
                assigned_personnel=user,
                activity_name=activity_name_to_match
            )

            # Combine descriptions from all relevant WARs
            description = " ".join([w.description for w in wars if w.description]) or ""

            # Collect WAR IDs to keep track
            war_ids = [w.id for w in wars]

            reports.append({
                "indicator": indicator.code,
                "description": description,
                "remarks": "COMPLIED" if description else "",
                "war_ids": war_ids,
            })

    context = {
        "reports": reports,
        "month_filter": month_filter,
        "unit_filter": unit_filter,
        "personnel_names": personnel_names,
    }

    return render(request, "gso_office/ipmt/ipmt_preview.html", context)

# -------------------------------
# Save IPMT
# -------------------------------
@login_required
@user_passes_test(is_gso_or_director)
def save_ipmt(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=400)

    try:
        data = json.loads(request.body)
        month = data.get("month")
        unit_name = data.get("unit")
        personnel_names = data.get("personnel", [])
        rows = data.get("rows", [])
    except Exception as e:
        return JsonResponse({"error": f"Invalid JSON: {str(e)}"}, status=400)

    unit = Unit.objects.filter(name__iexact=unit_name).first()
    if not unit:
        return JsonResponse({"error": "Unit not found"}, status=404)

    for person_name in personnel_names:
        user = get_user_by_identifier(person_name)
        if not user:
            continue

        for row in rows:
            indicator = SuccessIndicator.objects.filter(unit=unit, code=row.get("indicator")).first()
            if not indicator:
                # Optionally create indicator if missing
                indicator = SuccessIndicator.objects.create(
                    unit=unit,
                    code=row.get("indicator"),
                    name=row.get("indicator"),
                    is_active=True
                )

            # Fetch WARs for this indicator
            war_ids = row.get("war_ids", [])
            wars = WorkAccomplishmentReport.objects.filter(
                assigned_personnel=user,
                unit=unit,
                activity_name=indicator.code
            )
            if war_ids:
                wars = wars.filter(id__in=war_ids)

            # Determine accomplishment
            accomplishment = row.get("description", "").strip()
            remarks = row.get("remarks", "").strip() or accomplishment

            # Save or update IPMT
            ipmt_obj, created = IPMT.objects.update_or_create(
                personnel=user,
                unit=unit,
                month=month,
                indicator=indicator,
                defaults={
                    "accomplishment": accomplishment,
                    "remarks": remarks
                }
            )
            # Link WARs
            ipmt_obj.reports.set(wars)

    return JsonResponse({"status": "success"})




# --- Helper ---
def get_user_by_identifier(identifier):
    """Find user by username, full name, or partial name (case-insensitive)."""
    identifier = identifier.strip()
    if not identifier:
        return None

    # Try exact username
    user = User.objects.filter(username__iexact=identifier).first()
    if user:
        return user

    # Try full name match
    parts = identifier.split()
    if len(parts) >= 2:
        first, last = parts[0], parts[-1]
        user = User.objects.filter(
            Q(first_name__iexact=first) & Q(last_name__iexact=last)
        ).first()
        if user:
            return user

    # Try partial match
    return (
        User.objects.filter(
            Q(first_name__icontains=identifier) |
            Q(last_name__icontains=identifier)
        ).first()
    )